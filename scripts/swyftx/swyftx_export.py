#!/usr/bin/env python3
"""Export Swyftx transaction history + portfolio summary to a styled .xlsx.

Read-only. Reuses auth/data helpers from swyftx_client.py, so it needs
SWYFTX_API_KEY in the environment (run via `zsh -i -c` so ~/.zshrc is sourced).
Requires openpyxl (already installed system-wide).

Usage:
  python3 scripts/swyftx/swyftx_export.py [OUTPUT_PATH]

Default output: ~/Documents/Swyftx/swyftx-transactions-YYYY-MM-DD.xlsx
"""
import os
import sys
import time

import swyftx_client as sx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="0B5394")     # Swyftx-ish blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=13)   # larger header row (per pref)
TITLE_FONT = Font(bold=True, size=16, color="0B5394")
MONEY = "#,##0.00"


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def build(out_path):
    # --- pull data (read-only) ---
    hist = sx.cmd_history(["all"])["history"]
    port = sx.cmd_portfolio([])
    today = time.strftime("%Y-%m-%d")

    wb = Workbook()

    # ---- Sheet 1: Portfolio summary ----
    ws = wb.active
    ws.title = "Portfolio"
    ws["A1"] = "Swyftx Portfolio — %s (NZD)" % today
    ws["A1"].font = TITLE_FONT
    ws.append([])
    headers = ["Asset", "Name", "Balance", "Avg Cost (NZD)", "Price (NZD)",
               "Value (NZD)", "Price (USD)", "Value (USD)", "Invested (NZD)",
               "P/L (NZD)", "P/L %"]
    ws.append(headers)
    hdr_row = ws.max_row
    for h in port["holdings"]:
        ws.append([h["code"], h.get("name"), h["balance"], h.get("avg_cost_nzd"),
                   h.get("price_nzd"), h.get("value_nzd"), h.get("price_usd"),
                   h.get("value_usd"), h.get("invested_nzd"),
                   h.get("pl_nzd"), h.get("pl_pct")])
    t = port["totals"]
    ws.append([])
    ws.append(["TOTAL", "", "", "", "", t.get("value_nzd"), "", t.get("value_usd"),
               t.get("invested_nzd"), t.get("pl_nzd"), t.get("pl_pct")])
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    fx = port.get("usd_per_nzd")
    ws.append([])
    ws.append(["FX used: 1 NZD = %s USD" % fx if fx else ""])
    _style_header(ws, hdr_row)
    for col in ("D", "E", "F", "G", "H", "I", "J"):
        for cell in ws[col]:
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY
    _autosize(ws, [10, 22, 14, 15, 14, 13, 14, 13, 15, 13, 8])

    # ---- Sheet 2: Transactions ----
    ws2 = wb.create_sheet("Transactions")
    cols = ["Date (UTC)", "Type", "Status", "Asset", "Quantity",
            "Quantity Asset", "Primary Asset"]
    ws2.append(cols)
    for r in hist:
        ws2.append([r.get("date"), r.get("actionType"), r.get("status"),
                    r.get("asset"), r.get("quantity"), r.get("quantityAsset"),
                    r.get("primaryAsset")])
    _style_header(ws2, 1)
    for cell in ws2["E"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = MONEY
    _autosize(ws2, [20, 14, 12, 10, 14, 16, 14])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path, len(hist), len(port["holdings"])


def main(argv):
    default = os.path.expanduser(
        "~/Documents/Swyftx/swyftx-transactions-%s.xlsx" % time.strftime("%Y-%m-%d"))
    out_path = argv[0] if argv else default
    path, n_tx, n_hold = build(out_path)
    print("Wrote %s (%d transactions, %d holdings)" % (path, n_tx, n_hold))
    return 0


if __name__ == "__main__":
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    sys.exit(main(sys.argv[1:]))
